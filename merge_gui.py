# -*- coding: utf-8 -*-
import os
import glob                      # <--- 关键修复：导入 glob
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
import threading
from datetime import datetime

class ExcelMergeApp:
    def __init__(self, root):
        self.root = root
        root.title("兰湘子门店数据汇总工具")
        root.geometry("600x280")
        root.resizable(False, False)

        # 文件夹选择
        tk.Label(root, text="请选择包含 Excel 文件的文件夹：", font=("Arial", 10)).pack(pady=(20, 5))
        frame_folder = tk.Frame(root)
        frame_folder.pack(pady=5, padx=20, fill=tk.X)
        self.folder_var = tk.StringVar()
        self.folder_entry = tk.Entry(frame_folder, textvariable=self.folder_var, width=50)
        self.folder_entry.pack(side=tk.LEFT, padx=(0, 10), expand=True, fill=tk.X)
        self.folder_btn = tk.Button(frame_folder, text="浏览...", command=self.select_folder)
        self.folder_btn.pack(side=tk.RIGHT)

        # 输出文件路径（留空则自动生成到桌面，文件名包含日期）
        tk.Label(root, text="输出文件路径（留空则自动保存到桌面）：", font=("Arial", 10)).pack(pady=(10, 5))
        frame_output = tk.Frame(root)
        frame_output.pack(pady=5, padx=20, fill=tk.X)
        self.output_var = tk.StringVar()
        self.output_entry = tk.Entry(frame_output, textvariable=self.output_var, width=50)
        self.output_entry.pack(side=tk.LEFT, padx=(0, 10), expand=True, fill=tk.X)
        self.output_btn = tk.Button(frame_output, text="浏览...", command=self.select_output)
        self.output_btn.pack(side=tk.RIGHT)

        # 执行按钮
        self.run_btn = tk.Button(root, text="开始汇总", command=self.start_processing, bg="#4CAF50", fg="white", font=("Arial", 12), padx=20)
        self.run_btn.pack(pady=20)

        # 状态标签
        self.status_label = tk.Label(root, text="就绪", fg="blue", font=("Arial", 9))
        self.status_label.pack()

        # 进度条
        self.progress = ttk.Progressbar(root, mode='indeterminate', length=400)
        self.progress.pack(pady=10)

    def select_folder(self):
        folder = filedialog.askdirectory(title="选择存放 Excel 文件的文件夹")
        if folder:
            self.folder_var.set(folder)

    def select_output(self):
        try:
            file = filedialog.asksaveasfilename(
                title="保存汇总结果为",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            if file:
                self.output_var.set(file)
        except Exception as e:
            messagebox.showerror("错误", f"选择文件时出错：{e}\n请手动输入路径或留空使用默认路径。")

    def start_processing(self):
        folder = self.folder_var.get().strip()
        if not folder or not os.path.exists(folder):
            messagebox.showerror("错误", "请选择有效的文件夹！")
            return

        output_file = self.output_var.get().strip()
        if not output_file:
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            date_str = datetime.now().strftime("%Y%m%d")
            filename = f"数据汇总_{date_str}.xlsx"
            output_file = os.path.join(desktop, filename)
            self.output_var.set(output_file)

        self.run_btn.config(state=tk.DISABLED)
        self.status_label.config(text="正在处理...", fg="orange")
        self.progress.start()

        thread = threading.Thread(target=self.process_files, args=(folder, output_file))
        thread.daemon = True
        thread.start()

    def process_files(self, folder, output_file):
        try:
            self._run_processing(folder, output_file)
            self.root.after(0, self.on_success, output_file)
        except Exception as e:
            self.root.after(0, self.on_error, str(e))

    def _run_processing(self, folder, output_file):
        data = []  # 每个元素为 (store, material, remark)

        for file_path in glob.glob(os.path.join(folder, '*.xlsx')):
            try:
                wb = load_workbook(file_path, data_only=True)
                # 若需要固定读取 Sheet3，请将下一行改为 ws = wb['Sheet3']
                ws = wb.worksheets[0]
            except Exception:
                continue

            for row in ws.iter_rows(min_row=2, max_col=ws.max_column):
                store = row[0].value
                material = row[1].value
                remark = row[-1].value
                if remark is None or str(remark).strip() == '':
                    continue
                data.append((
                    str(store).strip() if store is not None else '',
                    str(material).strip() if material is not None else '',
                    str(remark).strip()
                ))

        if not data:
            raise ValueError("未提取到任何有效数据，请检查文件格式和内容。")

        # 去重并排序
        unique_data = list(set(data))
        unique_data.sort(key=lambda x: (x[0], x[1]))

        # 创建输出工作簿
        out_wb = Workbook()
        out_ws = out_wb.active
        out_ws.title = "汇总"

        # 写入标题
        out_ws['A1'] = '门店'
        out_ws['B1'] = '原料'
        out_ws['C1'] = '备注'

        # 写入数据
        for idx, (store, material, remark) in enumerate(unique_data, start=2):
            out_ws.cell(row=idx, column=1, value=store)
            out_ws.cell(row=idx, column=2, value=material)
            out_ws.cell(row=idx, column=3, value=remark)

        # 合并相同门店和相同原料的单元格（按顺序）
        self._merge_column(out_ws, col_idx=1, start_row=2)
        self._merge_column(out_ws, col_idx=2, start_row=2)

        # 设置列宽
        out_ws.column_dimensions['A'].width = 20
        out_ws.column_dimensions['B'].width = 30
        out_ws.column_dimensions['C'].width = 40

        out_wb.save(output_file)

    def _merge_column(self, ws, col_idx, start_row):
        current_val = None
        merge_start = None
        max_row = ws.max_row

        for row in range(start_row, max_row + 2):
            if row <= max_row:
                cell_val = ws.cell(row=row, column=col_idx).value
            else:
                cell_val = None

            if current_val is None:
                current_val = cell_val
                merge_start = row
            elif cell_val != current_val:
                if merge_start < row - 1:
                    ws.merge_cells(
                        start_row=merge_start, start_column=col_idx,
                        end_row=row - 1, end_column=col_idx
                    )
                    merged_cell = ws.cell(row=merge_start, column=col_idx)
                    merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                current_val = cell_val
                merge_start = row

    def on_success(self, output_file):
        self.progress.stop()
        self.run_btn.config(state=tk.NORMAL)
        self.status_label.config(text=f"✅ 汇总完成！已保存至：{output_file}", fg="green")
        messagebox.showinfo("完成", f"汇总表格已生成：\n{output_file}")
        try:
            os.startfile(os.path.dirname(output_file))
        except Exception:
            pass

    def on_error(self, error_msg):
        self.progress.stop()
        self.run_btn.config(state=tk.NORMAL)
        self.status_label.config(text=f"❌ 错误：{error_msg}", fg="red")
        messagebox.showerror("错误", f"处理出错：\n{error_msg}")

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelMergeApp(root)
    root.mainloop()